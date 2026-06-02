from __future__ import annotations

import shutil
import unittest
import uuid
import json
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

from project_tracker_backend import (
    PHOENIX_TASKS,
    ProjectRecord,
    ProjectTrackerBackend,
    parse_currency,
)
from user_auth import AuthStoreError, UserManager
from updater import (
    UpdatePackageError,
    _build_update_powershell_script,
    _validate_update_zip,
)


class TempWorkspaceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = Path(".test_tmp") / uuid.uuid4().hex
        self.tmp.mkdir(parents=True)
        self.addCleanup(lambda: shutil.rmtree(self.tmp, ignore_errors=True))


class AuthRegressionTests(TempWorkspaceTest):
    def test_corrupt_user_store_does_not_look_empty(self) -> None:
        users_path = self.tmp / "users.json"
        users_path.write_text("{bad json", encoding="utf-8")

        with self.assertRaises(AuthStoreError):
            UserManager(users_path)

    def test_first_created_user_is_admin(self) -> None:
        manager = UserManager(self.tmp / "users.json")
        manager.create_user("alice", "password123", role="user")

        user = manager.get_user("alice")
        assert user is not None
        self.assertEqual(user.role, "admin")

    def test_remember_me_requires_token_and_reset_invalidates_it(self) -> None:
        users_path = self.tmp / "users.json"
        manager = UserManager(users_path)
        manager.create_user("alice", "password123")

        token = manager.create_session_token("alice")
        user_data = json.loads(users_path.read_text(encoding="utf-8"))
        self.assertNotIn("session_token_hash", user_data["alice"])
        self.assertNotIn("session_token_expires_at", user_data["alice"])
        self.assertIsNotNone(manager.authenticate_session("alice", token))
        self.assertIsNone(manager.authenticate_session("alice", "not-the-token"))
        self.assertIsNone(manager.authenticate_session("alice", ""))

        manager.reset_password("alice", "temporary123")
        self.assertIsNone(manager.authenticate_session("alice", token))

    def test_remember_me_token_expires(self) -> None:
        users_path = self.tmp / "users.json"
        manager = UserManager(users_path)
        manager.create_user("alice", "password123")
        token = manager.create_session_token("alice")

        sessions_path = users_path.with_name("user_sessions.json")
        data = json.loads(sessions_path.read_text(encoding="utf-8"))
        data["alice"]["expires_at"] = (
            datetime.now() - timedelta(days=1)
        ).replace(microsecond=0).isoformat(sep=" ")
        sessions_path.write_text(json.dumps(data), encoding="utf-8")

        self.assertIsNone(manager.authenticate_session("alice", token))

    def test_v180_session_fields_are_migrated_out_of_users_file(self) -> None:
        users_path = self.tmp / "users.json"
        token = "saved-token"
        expires_at = (datetime.now() + timedelta(days=1)).replace(
            microsecond=0
        ).isoformat(sep=" ")
        users_path.write_text(
            json.dumps(
                {
                    "alice": {
                        "username": "alice",
                        "password_hash": "hash",
                        "salt": "00",
                        "must_change_password": False,
                        "created_at": "",
                        "role": "admin",
                        "session_token_hash": token,
                        "session_token_expires_at": expires_at,
                    }
                }
            ),
            encoding="utf-8",
        )

        UserManager(users_path)

        users_data = json.loads(users_path.read_text(encoding="utf-8"))
        self.assertNotIn("session_token_hash", users_data["alice"])
        self.assertNotIn("session_token_expires_at", users_data["alice"])
        sessions_data = json.loads(
            users_path.with_name("user_sessions.json").read_text(encoding="utf-8")
        )
        self.assertEqual(sessions_data["alice"]["token_hash"], token)


class BackendRegressionTests(TempWorkspaceTest):
    def test_parse_currency_handles_user_formatted_values(self) -> None:
        self.assertEqual(parse_currency("$1,234.56"), 1234.56)
        self.assertEqual(parse_currency("(1,234.56)"), -1234.56)
        self.assertEqual(parse_currency("not money"), 0.0)

    def test_excel_export_accepts_formatted_contract_value(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        project_id = backend.create_project(
            ProjectRecord(
                job_name="Formatted Contract",
                job_number="12345",
                contract_value="$1,234.56",
            )
        )

        output_path = backend.export_project_to_excel(project_id, self.tmp / "out.xlsx")
        self.assertTrue(output_path.exists())

    def test_dashboard_task_counts_exclude_test_jobs(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        backend.create_project(ProjectRecord(job_name="Real", job_number="R"))
        backend.create_project(ProjectRecord(job_name="Training", job_number="T", is_test=True))

        stats = backend.get_dashboard_stats()

        self.assertEqual(stats["project_count"], 1)
        self.assertEqual(stats["total_tasks"], len(PHOENIX_TASKS))
        self.assertEqual(stats["incomplete_count"], len(PHOENIX_TASKS))

    def test_new_projects_default_to_phoenix_template(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        project_id = backend.create_project(ProjectRecord(job_name="Job", job_number="J"))

        expected_names = {task["task_name"] for task in PHOENIX_TASKS}
        actual_names = {task.task_name for task in backend.list_tasks(project_id)}

        self.assertEqual(actual_names, expected_names)

    def test_deleting_task_removes_task_notes(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        project_id = backend.create_project(
            ProjectRecord(job_name="Job", job_number="J"),
            task_template="standard",
        )
        task = backend.list_tasks(project_id)[0]
        task_id = task.id
        assert task_id is not None

        backend.add_task_note(task_id, "note")
        backend.delete_task(task_id)

        self.assertEqual(backend.list_task_notes(task_id), [])

    def test_replacing_tasks_removes_blank_tasks_not_in_target_template(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        project_id = backend.create_project(
            ProjectRecord(job_name="Job", job_number="J"),
            task_template="standard",
        )

        backend.replace_project_tasks(project_id, "phoenix")

        phoenix_names = {task["task_name"] for task in PHOENIX_TASKS}
        actual_names = {task.task_name for task in backend.list_tasks(project_id)}
        self.assertTrue(actual_names.issubset(phoenix_names))

    def test_replacing_tasks_preserves_completed_and_noted_tasks(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        project_id = backend.create_project(
            ProjectRecord(job_name="Job", job_number="J"),
            task_template="standard",
        )
        noted_excluded_task = next(
            task for task in backend.list_tasks(project_id)
            if task.task_name == "Job Plan Developed"
        )
        completed_excluded_task = next(
            task for task in backend.list_tasks(project_id)
            if task.task_name == "Service Turnover"
        )
        shared_task = next(
            task for task in backend.list_tasks(project_id)
            if task.task_name == "Phoenix Material Submittal"
        )
        assert noted_excluded_task.id is not None
        assert completed_excluded_task.id is not None
        assert shared_task.id is not None

        backend.add_task_note(noted_excluded_task.id, "kept history")
        backend.set_task_completed(completed_excluded_task.id, True, "2026-04-01")
        backend.update_task(shared_task.id, notes="kept inline note")

        backend.replace_project_tasks(project_id, "phoenix")
        tasks_after_phoenix = {task.task_name: task for task in backend.list_tasks(project_id)}

        self.assertIn("Job Plan Developed", tasks_after_phoenix)
        self.assertIn("Service Turnover", tasks_after_phoenix)
        self.assertTrue(tasks_after_phoenix["Service Turnover"].is_complete)
        self.assertEqual(tasks_after_phoenix["Service Turnover"].completed_date, "2026-04-01")
        self.assertEqual(
            [note.content for note in backend.list_task_notes(noted_excluded_task.id)],
            ["kept history"],
        )
        self.assertEqual(
            tasks_after_phoenix["Phoenix Material Submittal"].notes,
            "kept inline note",
        )

        backend.replace_project_tasks(project_id, "standard")
        tasks_after_standard = {task.task_name: task for task in backend.list_tasks(project_id)}

        self.assertTrue(tasks_after_standard["Service Turnover"].is_complete)
        self.assertEqual(tasks_after_standard["Service Turnover"].completed_date, "2026-04-01")
        self.assertEqual(
            [note.content for note in backend.list_task_notes(noted_excluded_task.id)],
            ["kept history"],
        )
        self.assertEqual(
            tasks_after_standard["Phoenix Material Submittal"].notes,
            "kept inline note",
        )


class UpdaterRegressionTests(TempWorkspaceTest):
    def test_update_zip_requires_internal_runtime_folder(self) -> None:
        zip_path = self.tmp / "ProjectTrackingTool.zip"
        with zipfile.ZipFile(zip_path, "w") as zf:
            zf.writestr("ProjectTrackingTool.exe", "stub")

        with self.assertRaises(UpdatePackageError):
            _validate_update_zip(zip_path)

    def test_update_zip_accepts_full_flat_payload(self) -> None:
        zip_path = self.tmp / "ProjectTrackingTool.zip"
        with zipfile.ZipFile(zip_path, "w") as zf:
            zf.writestr("ProjectTrackingTool.exe", "stub")
            zf.writestr("_internal/runtime.dll", "stub")

        _validate_update_zip(zip_path)

    def test_update_script_copies_payload_contents_to_install_folder(self) -> None:
        script = _build_update_powershell_script(
            self.tmp / "ProjectTrackingTool.zip",
            self.tmp / "install",
            self.tmp / "install" / "ProjectTrackingTool.exe",
        )

        self.assertIn("Get-ChildItem -LiteralPath $payload -Force", script)
        self.assertIn("Copy-Item -Destination $installDir -Recurse -Force", script)
        self.assertIn("_internal", script)
        self.assertIn("Write-Output \"Starting update from $zipPath\"", script)
        self.assertIn("Write-Output \"Update files copied successfully.\"", script)
        self.assertNotIn("Out-File -FilePath $logPath", script)


class V185RegressionTests(TempWorkspaceTest):
    """Regression tests covering the v1.8.5 bug-fix release."""

    # ── L2: timezone-aware session expiry ────────────────────────────────
    def test_session_expiry_treats_legacy_naive_timestamp_as_valid(self) -> None:
        users_path = self.tmp / "users.json"
        manager = UserManager(users_path)
        manager.create_user("alice", "password123")
        token = manager.create_session_token("alice")

        sessions_path = users_path.with_name("user_sessions.json")
        data = json.loads(sessions_path.read_text(encoding="utf-8"))
        future_naive = (datetime.now() + timedelta(days=10)).replace(
            microsecond=0
        ).isoformat(sep=" ")
        data["alice"]["expires_at"] = future_naive
        sessions_path.write_text(json.dumps(data), encoding="utf-8")

        self.assertIsNotNone(manager.authenticate_session("alice", token))

    def test_session_expiry_written_with_timezone(self) -> None:
        users_path = self.tmp / "users.json"
        manager = UserManager(users_path)
        manager.create_user("alice", "password123")
        manager.create_session_token("alice")

        sessions_path = users_path.with_name("user_sessions.json")
        data = json.loads(sessions_path.read_text(encoding="utf-8"))
        self.assertIn("+00:00", data["alice"]["expires_at"])

    # ── M2: reset_password is atomic ─────────────────────────────────────
    def test_reset_password_sets_must_change_flag_and_new_password(self) -> None:
        manager = UserManager(self.tmp / "users.json")
        manager.create_user("alice", "password123", must_change_password=False)
        manager.reset_password("alice", "temporary123")

        user = manager.get_user("alice")
        assert user is not None
        self.assertTrue(user.must_change_password)
        self.assertIsNotNone(manager.authenticate("alice", "temporary123"))

    def test_reset_password_with_short_password_does_not_modify_account(self) -> None:
        manager = UserManager(self.tmp / "users.json")
        manager.create_user("alice", "password123")
        with self.assertRaises(ValueError):
            manager.reset_password("alice", "short")
        self.assertIsNotNone(manager.authenticate("alice", "password123"))

    # ── L5: _parse_version returns None on unparseable ───────────────────
    def test_parse_version_returns_none_on_unparseable(self) -> None:
        from updater import _parse_version
        self.assertIsNone(_parse_version(""))
        self.assertIsNone(_parse_version("not-a-version"))
        self.assertIsNone(_parse_version("v1.2.beta"))

    def test_parse_version_handles_normal_tags(self) -> None:
        from updater import _parse_version
        self.assertEqual(_parse_version("v1.8.5"), (1, 8, 5))
        self.assertEqual(_parse_version("1.8.5"), (1, 8, 5))
        self.assertEqual(_parse_version("V2.0.0"), (2, 0, 0))

    # ── L6: job_number uniqueness is case-insensitive ────────────────────
    def test_create_project_rejects_case_different_duplicate_job_number(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        backend.create_project(ProjectRecord(job_name="A", job_number="ABC-123"))
        with self.assertRaises(ValueError):
            backend.create_project(ProjectRecord(job_name="B", job_number="abc-123"))

    def test_update_project_rejects_case_different_duplicate_job_number(self) -> None:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        backend.create_project(ProjectRecord(job_name="A", job_number="ABC-123"))
        pid2 = backend.create_project(ProjectRecord(job_name="B", job_number="XYZ-999"))
        with self.assertRaises(ValueError):
            backend.update_project(pid2, job_number="abc-123")

    # ── M3: cache invalidated on save failure ────────────────────────────
    def test_cache_invalidated_after_save_failure(self) -> None:
        import tempfile as _tempfile
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        backend.create_project(ProjectRecord(job_name="Original", job_number="1"))

        real_mkstemp = _tempfile.mkstemp
        def failing_mkstemp(*args, **kwargs):
            raise OSError("simulated disk failure")
        _tempfile.mkstemp = failing_mkstemp  # type: ignore[assignment]
        # The backend module imported tempfile at the top, so we have to patch
        # the symbol it actually uses.
        import project_tracker_backend as _ptb
        _ptb.tempfile.mkstemp = failing_mkstemp  # type: ignore[attr-defined]
        try:
            with self.assertRaises(OSError):
                backend.create_project(ProjectRecord(job_name="Failed", job_number="2"))
        finally:
            _tempfile.mkstemp = real_mkstemp  # type: ignore[assignment]
            _ptb.tempfile.mkstemp = real_mkstemp  # type: ignore[attr-defined]

        # The failed mutation must not be visible on the next read.
        projects = backend.list_projects()
        self.assertEqual(len(projects), 1)
        self.assertEqual(projects[0].job_name, "Original")

    # ── H3: _project_from_dict tolerates missing fields ──────────────────
    def test_load_tolerates_project_record_missing_optional_fields(self) -> None:
        data_path = self.tmp / "data.json"
        data_path.write_text(json.dumps({
            "projects": [{"id": 1, "job_name": "Sparse"}],
            "tasks": [], "notes": [], "change_orders": [], "activity_log": [],
            "task_notes": [], "address_book": [], "pending_deletes": [],
            "next_project_id": 2, "next_task_id": 1, "next_note_id": 1,
            "next_co_id": 1, "next_activity_id": 1, "next_task_note_id": 1,
            "next_address_id": 1, "next_pending_id": 1,
        }), encoding="utf-8")

        backend = ProjectTrackerBackend(data_path)
        projects = backend.list_projects()
        self.assertEqual(len(projects), 1)
        self.assertEqual(projects[0].job_name, "Sparse")
        self.assertEqual(projects[0].job_number, "")
        self.assertEqual(projects[0].project_manager, "")

    # ── H4: workbook validation ──────────────────────────────────────────
    def test_import_workbook_rejects_unrelated_xlsx(self) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Date"
        ws["B1"] = "Amount"
        ws["A2"] = "2026-01-01"
        ws["B2"] = 100
        bad_path = self.tmp / "expenses.xlsx"
        wb.save(bad_path)

        backend = ProjectTrackerBackend(self.tmp / "data.json")
        with self.assertRaises(ValueError):
            backend.import_project_from_workbook(bad_path)

    def test_import_workbook_accepts_phoenix_header(self) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["C3"] = "Test Project"
        ws["H3"] = "12345"
        ws["B10"] = "Sales-Ops Turnover"
        good_path = self.tmp / "phoenix.xlsx"
        wb.save(good_path)

        backend = ProjectTrackerBackend(self.tmp / "data.json")
        project_id = backend.import_project_from_workbook(good_path)
        project = backend.get_project(project_id)
        assert project is not None
        self.assertEqual(project.job_name, "Test Project")
        self.assertEqual(project.job_number, "12345")

    # ── L1: backup error surfacing ───────────────────────────────────────
    def test_backup_returns_none_on_success(self) -> None:
        from project_tracker_gui import _backup_data_file
        data_path = self.tmp / "data.json"
        data_path.write_text("{}", encoding="utf-8")

        result = _backup_data_file(data_path)
        self.assertIsNone(result)
        backups = list((data_path.parent / "backups").glob("*.json"))
        self.assertEqual(len(backups), 1)

    def test_backup_returns_none_when_source_file_missing(self) -> None:
        from project_tracker_gui import _backup_data_file
        result = _backup_data_file(self.tmp / "no_such_file.json")
        self.assertIsNone(result)


class JTF2WebProIDsTests(TempWorkspaceTest):
    """Backend data-compatibility behavior for JTF-2 multi-WebPro support."""

    def _make_backend(self) -> ProjectTrackerBackend:
        return ProjectTrackerBackend(self.tmp / "data.json")

    # ── Read-path migration / fallback ────────────────────────────────────
    def test_legacy_single_webpro_id_loads_as_single_item_list(self) -> None:
        data_path = self.tmp / "data.json"
        data_path.write_text(json.dumps({
            "projects": [{
                "id": 1, "job_name": "Legacy", "job_number": "L-1",
                "webpro_id": "12345",  # legacy field only, no webpro_ids
            }],
            "tasks": [], "notes": [], "change_orders": [], "activity_log": [],
            "task_notes": [], "address_book": [], "pending_deletes": [],
            "next_project_id": 2, "next_task_id": 1, "next_note_id": 1,
            "next_co_id": 1, "next_activity_id": 1, "next_task_note_id": 1,
            "next_address_id": 1, "next_pending_id": 1,
        }), encoding="utf-8")

        backend = ProjectTrackerBackend(data_path)
        project = backend.get_project(1)
        assert project is not None
        self.assertEqual(project.webpro_ids, ["12345"])
        self.assertEqual(project.webpro_id, "12345")  # legacy mirror preserved

    def test_new_webpro_ids_list_loads_directly(self) -> None:
        data_path = self.tmp / "data.json"
        data_path.write_text(json.dumps({
            "projects": [{
                "id": 1, "job_name": "Modern", "job_number": "M-1",
                "webpro_id": "abc",
                "webpro_ids": ["abc", "def", "xyz"],
            }],
            "tasks": [], "notes": [], "change_orders": [], "activity_log": [],
            "task_notes": [], "address_book": [], "pending_deletes": [],
            "next_project_id": 2, "next_task_id": 1, "next_note_id": 1,
            "next_co_id": 1, "next_activity_id": 1, "next_task_note_id": 1,
            "next_address_id": 1, "next_pending_id": 1,
        }), encoding="utf-8")

        backend = ProjectTrackerBackend(data_path)
        project = backend.get_project(1)
        assert project is not None
        self.assertEqual(project.webpro_ids, ["abc", "def", "xyz"])
        self.assertEqual(project.webpro_id, "abc")  # mirrors first item

    def test_missing_both_webpro_fields_loads_as_empty_list(self) -> None:
        data_path = self.tmp / "data.json"
        data_path.write_text(json.dumps({
            "projects": [{"id": 1, "job_name": "Bare", "job_number": "B-1"}],
            "tasks": [], "notes": [], "change_orders": [], "activity_log": [],
            "task_notes": [], "address_book": [], "pending_deletes": [],
            "next_project_id": 2, "next_task_id": 1, "next_note_id": 1,
            "next_co_id": 1, "next_activity_id": 1, "next_task_note_id": 1,
            "next_address_id": 1, "next_pending_id": 1,
        }), encoding="utf-8")

        backend = ProjectTrackerBackend(data_path)
        project = backend.get_project(1)
        assert project is not None
        self.assertEqual(project.webpro_ids, [])
        self.assertEqual(project.webpro_id, "")

    # ── Write-path dual-write + normalization ─────────────────────────────
    def test_create_project_writes_both_webpro_id_and_webpro_ids(self) -> None:
        backend = self._make_backend()
        pid = backend.create_project(
            ProjectRecord(
                job_name="Dual",
                job_number="D-1",
                webpro_ids=["alpha", "beta"],
            )
        )

        raw = json.loads((self.tmp / "data.json").read_text(encoding="utf-8"))
        stored = next(p for p in raw["projects"] if int(p["id"]) == pid)
        self.assertEqual(stored["webpro_ids"], ["alpha", "beta"])
        self.assertEqual(stored["webpro_id"], "alpha")  # legacy mirror = first

    def test_create_project_with_legacy_single_field_still_works(self) -> None:
        backend = self._make_backend()
        # Older callers may construct a ProjectRecord with only webpro_id.
        pid = backend.create_project(
            ProjectRecord(job_name="Legacy caller", job_number="L-1", webpro_id="solo")
        )
        project = backend.get_project(pid)
        assert project is not None
        self.assertEqual(project.webpro_ids, ["solo"])
        self.assertEqual(project.webpro_id, "solo")

    def test_update_project_dedupes_case_insensitive_preserves_order(self) -> None:
        backend = self._make_backend()
        pid = backend.create_project(ProjectRecord(job_name="X", job_number="X-1"))
        backend.update_project(pid, webpro_ids=["abc", "ABC", "Def", "abc", "def"])
        project = backend.get_project(pid)
        assert project is not None
        # First-seen casing wins; duplicates dropped; insertion order preserved.
        self.assertEqual(project.webpro_ids, ["abc", "Def"])
        self.assertEqual(project.webpro_id, "abc")

    def test_update_project_strips_whitespace_and_drops_empty(self) -> None:
        backend = self._make_backend()
        pid = backend.create_project(ProjectRecord(job_name="Y", job_number="Y-1"))
        backend.update_project(pid, webpro_ids=["  abc  ", "", "   ", "def\t"])
        project = backend.get_project(pid)
        assert project is not None
        self.assertEqual(project.webpro_ids, ["abc", "def"])

    def test_update_project_via_legacy_webpro_id_still_works(self) -> None:
        backend = self._make_backend()
        pid = backend.create_project(ProjectRecord(job_name="Z", job_number="Z-1"))
        backend.update_project(pid, webpro_id="hello")
        project = backend.get_project(pid)
        assert project is not None
        self.assertEqual(project.webpro_ids, ["hello"])

    def test_existing_app_reading_new_file_sees_first_id_as_webpro_id(self) -> None:
        """Forward-compat: an older app version reading a v1.8.7+ file should
        still see the legacy ``webpro_id`` key populated with the first ID,
        so its single-string display code keeps working."""
        backend = self._make_backend()
        pid = backend.create_project(
            ProjectRecord(job_name="Forward", job_number="F-1", webpro_ids=["first", "second"])
        )
        raw = json.loads((self.tmp / "data.json").read_text(encoding="utf-8"))
        stored = next(p for p in raw["projects"] if int(p["id"]) == pid)
        self.assertIn("webpro_id", stored)
        self.assertEqual(stored["webpro_id"], "first")

    def test_clearing_webpro_ids_removes_legacy_value_too(self) -> None:
        backend = self._make_backend()
        pid = backend.create_project(
            ProjectRecord(job_name="Clearable", job_number="C-1", webpro_ids=["one", "two"])
        )
        backend.update_project(pid, webpro_ids=[])
        project = backend.get_project(pid)
        assert project is not None
        self.assertEqual(project.webpro_ids, [])
        self.assertEqual(project.webpro_id, "")

    # ── Search inclusion ──────────────────────────────────────────────────
    def test_list_projects_search_matches_against_any_webpro_id(self) -> None:
        backend = self._make_backend()
        backend.create_project(
            ProjectRecord(job_name="Searchable", job_number="S-1",
                          webpro_ids=["needle-001", "decoy"])
        )
        backend.create_project(ProjectRecord(job_name="Other", job_number="O-1"))

        matches = backend.list_projects(search_text="needle")
        self.assertEqual([p.job_name for p in matches], ["Searchable"])

        matches_decoy = backend.list_projects(search_text="decoy")
        self.assertEqual([p.job_name for p in matches_decoy], ["Searchable"])

    def test_list_projects_search_also_matches_legacy_webpro_id_only(self) -> None:
        """If a stored record carries only the legacy webpro_id field,
        search should still find it via the migration-aware haystack."""
        data_path = self.tmp / "data.json"
        data_path.write_text(json.dumps({
            "projects": [{
                "id": 1, "job_name": "Legacy Search", "job_number": "LS-1",
                "webpro_id": "only-legacy",
            }],
            "tasks": [], "notes": [], "change_orders": [], "activity_log": [],
            "task_notes": [], "address_book": [], "pending_deletes": [],
            "next_project_id": 2, "next_task_id": 1, "next_note_id": 1,
            "next_co_id": 1, "next_activity_id": 1, "next_task_note_id": 1,
            "next_address_id": 1, "next_pending_id": 1,
        }), encoding="utf-8")
        backend = ProjectTrackerBackend(data_path)
        matches = backend.list_projects(search_text="only-legacy")
        self.assertEqual([p.job_name for p in matches], ["Legacy Search"])


class JTF1RSSFilterTests(TempWorkspaceTest):
    """Backend filter behavior for the JTF-1 RSS filter dropdown."""

    def _setup_three_projects(self) -> ProjectTrackerBackend:
        backend = ProjectTrackerBackend(self.tmp / "data.json")
        backend.create_project(ProjectRecord(job_name="Alpha", job_number="A-1"))
        backend.create_project(ProjectRecord(job_name="Beta", job_number="B-2"))
        backend.create_project(ProjectRecord(job_name="Gamma", job_number="G-3"))
        # Attach an RSS entry to Beta and Gamma by editing the stored list directly.
        beta = next(p for p in backend.list_projects() if p.job_number == "B-2")
        gamma = next(p for p in backend.list_projects() if p.job_number == "G-3")
        assert beta.id is not None and gamma.id is not None
        backend.update_project(
            beta.id,
            rss_files=[{"name": "feed", "path": "x.csv", "rows": []}],
        )
        backend.update_project(
            gamma.id,
            rss_files=[{"name": "other", "path": "y.csv", "rows": []}],
        )
        return backend

    def test_has_rss_true_returns_only_projects_with_rss(self) -> None:
        backend = self._setup_three_projects()
        with_rss = backend.list_projects(has_rss=True)
        names = sorted(p.job_name for p in with_rss)
        self.assertEqual(names, ["Beta", "Gamma"])

    def test_has_rss_false_returns_only_projects_without_rss(self) -> None:
        backend = self._setup_three_projects()
        without_rss = backend.list_projects(has_rss=False)
        names = sorted(p.job_name for p in without_rss)
        self.assertEqual(names, ["Alpha"])

    def test_has_rss_none_returns_all_projects(self) -> None:
        backend = self._setup_three_projects()
        all_projects = backend.list_projects(has_rss=None)
        names = sorted(p.job_name for p in all_projects)
        self.assertEqual(names, ["Alpha", "Beta", "Gamma"])

    def test_has_rss_filter_composes_with_text_search(self) -> None:
        backend = self._setup_three_projects()
        # Text matches Beta + Gamma's first letters; RSS filter further matches.
        matches = backend.list_projects(search_text="B", has_rss=True)
        names = sorted(p.job_name for p in matches)
        self.assertEqual(names, ["Beta"])

        # Text search alone (no RSS filter) returns Beta only because "B"
        # matches the job number prefix; sanity-check no contamination.
        matches_all = backend.list_projects(search_text="B")
        names_all = sorted(p.job_name for p in matches_all)
        self.assertEqual(names_all, ["Beta"])


if __name__ == "__main__":
    unittest.main()
