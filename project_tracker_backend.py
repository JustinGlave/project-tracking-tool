from __future__ import annotations

import json
import logging
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

DEFAULT_TEMPLATE_NAME = "Phoenix Job Tracking"


@dataclass(slots=True)
class ProjectRecord:
    id: Optional[int] = None
    job_name: str = ""
    job_number: str = ""
    project_manager: str = ""
    sales_engineer: str = ""
    target_completion: Optional[str] = None
    liquid_damages: str = ""
    warranty_period: str = ""
    notes: str = ""
    # Fields populated from Odin assignment email
    booked_date: str = ""
    group_ops_manager: str = ""
    group_ops_supervisor: str = ""
    job_subtype: str = ""
    owner: str = ""
    contracted_with: str = ""
    general_contractor: str = ""
    contract_value: str = ""
    job_docs: str = ""
    div25_url: str = ""
    webpro_id: str = ""
    is_test: bool = False
    pinned: bool = False
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    created_by: str = ""
    updated_by: str = ""
    rss_files: list = field(default_factory=list)


@dataclass(slots=True)
class TaskRecord:
    id: Optional[int] = None
    project_id: Optional[int] = None
    task_name: str = ""
    phase: str = "General"
    sort_order: int = 0
    due_date: Optional[str] = None
    completed_date: Optional[str] = None
    is_complete: bool = False
    notes: str = ""
    updated_by: str = ""


@dataclass(slots=True)
class ActivityRecord:
    id: Optional[int] = None
    project_id: Optional[int] = None
    timestamp: str = ""
    user: str = ""
    action: str = ""       # "created" | "updated" | "deleted" | "completed" | "uncompleted"
    entity_type: str = ""  # "task" | "project"
    entity_name: str = ""
    details: str = ""


@dataclass(slots=True)
class TaskNoteRecord:
    id: Optional[int] = None
    task_id: Optional[int] = None
    timestamp: str = ""
    user: str = ""
    content: str = ""


@dataclass(slots=True)
class NoteRecord:
    id: Optional[int] = None
    project_id: Optional[int] = None
    note_number: int = 0
    date: str = ""
    content: str = ""
    status: str = "Open"          # "Open" or "Closed"
    closeout_comment: str = ""


@dataclass(slots=True)
class ChangeOrderRecord:
    id: Optional[int] = None
    project_id: Optional[int] = None
    cop_number: str = ""
    reference: str = ""
    description: str = ""
    creation_date: str = ""
    ats_price: str = ""
    ats_direct_cost: str = ""
    ats_status: str = "Pending"       # Pending / Accepted / Rejected
    booked_in_portal: str = ""
    ats_booked_co: str = ""
    mech_co: str = ""
    sub_quoted_price: str = ""
    sub_plug_number: str = ""
    sub_status: str = "Pending"       # Pending / Accepted / Rejected
    sub_co_sent: str = ""
    sub_co_number: str = ""
    notes: str = ""


# Tuple of dicts — immutable at the container level, preventing accidental
# runtime mutation of the default task list.
DEFAULT_TASKS: tuple[dict[str, Any], ...] = (
    {"phase": "Pre-Project", "task_name": "Sales-Ops Turnover"},
    {"phase": "Pre-Project", "task_name": "Review of Specs/Add."},
    {"phase": "Pre-Project", "task_name": "Contract Review"},
    {"phase": "Planning", "task_name": "Job Plan Developed"},
    {"phase": "Materials", "task_name": "Phoenix Material Submittal"},
    {"phase": "Materials", "task_name": "Phoenix Material Submittal Approved"},
    {"phase": "Materials", "task_name": "Phoenix Material Delivery Confirmations"},
    {"phase": "Materials", "task_name": "Valves Ordered"},
    {"phase": "Shipping", "task_name": "Flow Curves Archived in Job Folder"},
    {"phase": "Shipping", "task_name": "Shipping Tracking to Customer"},
    {"phase": "Shipping", "task_name": "Shipping confirmed to Customer/ATS"},
    {"phase": "Engineering", "task_name": "Phoenix Drawing Package Submittal"},
    {"phase": "Engineering", "task_name": "PM Book, and Checkout Sheet Book Developed"},
    {"phase": "Engineering", "task_name": "Engineering Re-estimate"},
    {"phase": "Installation", "task_name": "Hire Elec. Sub Request Form"},
    {"phase": "Installation", "task_name": "Elec Install Standards reviewed with installer"},
    {"phase": "Controls", "task_name": "DDC Developed"},
    {"phase": "Controls", "task_name": "Graphics Developed"},
    {"phase": "Turnover", "task_name": "Service Turnover"},
    {"phase": "Commissioning", "task_name": "Submitted Commissioning Documents"},
    {"phase": "Commissioning", "task_name": "Checkout/PTP w/ Documentation"},
    {"phase": "Commissioning", "task_name": "Start-Up Complete All docs to CX and Ops Teams"},
    {"phase": "Commissioning", "task_name": "Commissioning Complete"},
    {"phase": "Closeout", "task_name": "All Punch Lists Complete"},
    {"phase": "Closeout", "task_name": "Close-out Documents Complete"},
    {"phase": "Closeout", "task_name": "Owner Training"},
    {"phase": "Archive", "task_name": "Job Back-up Archived"},
    {"phase": "Archive", "task_name": "As-Built Drawings Completed"},
    {"phase": "Archive", "task_name": "Archive Drawings (PDF)"},
    {"phase": "Archive", "task_name": "Scan Check-out Sheets to Job File"},
    {"phase": "Warranty", "task_name": "Warranty Letter"},
    {"phase": "Materials", "task_name": "Return Excess Materials"},
    {"phase": "Financial", "task_name": "Back-up Database Set-up Dial-up Database Tracking Form"},
    {"phase": "Financial", "task_name": "Resolve Trailing Costs"},
    {"phase": "Financial", "task_name": "Resolve All Change-orders"},
    {"phase": "Archive", "task_name": "Send record drawings to service Account Manager"},
    {"phase": "Financial", "task_name": "Final Billing and Payment"},
)

_PHOENIX_EXCLUDED: frozenset[str] = frozenset({
    "Job Plan Developed",
    "PM Book, and Checkout Sheet Book Developed",
    "Engineering Re-estimate",
    "Service Turnover",
    "Job Back-up Archived",
    "Scan Check-out Sheets to Job File",
    "Return Excess Materials",
    "Back-up Database Set-up Dial-up Database Tracking Form",
    "Resolve Trailing Costs",
    "Send record drawings to service Account Manager",
})

PHOENIX_TASKS: tuple[dict[str, Any], ...] = tuple(
    t for t in DEFAULT_TASKS if t["task_name"] not in _PHOENIX_EXCLUDED
)


def _migrate_rss_files(project_dict: dict) -> list:
    """Return rss_files list, migrating from legacy csv_file_path if needed."""
    rss_files = project_dict.get("rss_files")
    if isinstance(rss_files, list):
        return rss_files
    old_path = project_dict.get("csv_file_path", "")
    if old_path:
        return [{"name": "Imported", "path": old_path}]
    return []


class ProjectTrackerBackend:
    """JSON-backed project tracker service.

    Storage notes
    -------------
    All data lives in a single JSON file. This is intentionally simple and
    portable for a single-user desktop tool. There is no file-level locking,
    so running two instances against the same file simultaneously can produce
    lost updates. If multi-process access ever becomes a requirement, replace
    the load/save calls with a proper database (SQLite with WAL mode is the
    natural next step).

    Workbook import notes
    ---------------------
    ``import_project_from_workbook`` reads header fields from hard-coded cell
    addresses (C3, H3, C4, …) that match the Phoenix Job Tracking template
    layout. If the template is ever restructured those addresses will need
    updating here.
    """

    def __init__(self, db_path: str | Path = "project_tracker_data.json") -> None:
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.current_user: str = ""
        self._initialize_storage()

    def _initialize_storage(self) -> None:
        if self.db_path.exists():
            data = self._load_data()
            changed = False
            if "projects" not in data:
                data["projects"] = []; changed = True
            if "tasks" not in data:
                data["tasks"] = []; changed = True
            if "notes" not in data:
                data["notes"] = []; changed = True
            if "change_orders" not in data:
                data["change_orders"] = []; changed = True
            if "next_project_id" not in data:
                data["next_project_id"] = self._next_id(data["projects"]); changed = True
            if "next_task_id" not in data:
                data["next_task_id"] = self._next_id(data["tasks"]); changed = True
            if "next_note_id" not in data:
                data["next_note_id"] = self._next_id(data.get("notes", [])); changed = True
            if "next_co_id" not in data:
                data["next_co_id"] = self._next_id(data.get("change_orders", [])); changed = True
            if "activity_log" not in data:
                data["activity_log"] = []; changed = True
            if "next_activity_id" not in data:
                data["next_activity_id"] = 1; changed = True
            if "task_notes" not in data:
                data["task_notes"] = []; changed = True
            if "next_task_note_id" not in data:
                data["next_task_note_id"] = 1; changed = True
            if changed:
                self._save_data(data)
            return

        self._save_data({
            "projects": [], "tasks": [], "notes": [], "change_orders": [],
            "activity_log": [], "task_notes": [],
            "next_project_id": 1, "next_task_id": 1, "next_note_id": 1,
            "next_co_id": 1, "next_activity_id": 1, "next_task_note_id": 1,
        })

    def _load_data(self) -> dict[str, Any]:
        if not self.db_path.exists():
            return {
                "projects": [], "tasks": [], "notes": [], "change_orders": [],
                "activity_log": [], "task_notes": [],
                "next_project_id": 1, "next_task_id": 1, "next_note_id": 1,
                "next_co_id": 1, "next_activity_id": 1, "next_task_note_id": 1,
            }
        return json.loads(self.db_path.read_text(encoding="utf-8"))

    def _save_data(self, data: dict[str, Any]) -> None:
        # Write to a temp file alongside the target, then rename — this ensures
        # atomic saves on all major platforms and prevents a half-written
        # JSON file if the process is interrupted mid-save.
        import time
        tmp_fd, tmp_path_str = tempfile.mkstemp(
            dir=self.db_path.parent, suffix=".tmp"
        )
        tmp_path = Path(tmp_path_str)
        try:
            with open(tmp_fd, "w", encoding="utf-8") as fh:
                json.dump(data, fh, indent=2)
            # Retry rename up to 5 times — cloud-sync clients (OneDrive, etc.)
            # can briefly lock the target file, causing WinError 5.
            for attempt in range(5):
                try:
                    tmp_path.replace(self.db_path)
                    break
                except PermissionError:
                    if attempt == 4:
                        raise
                    time.sleep(0.2 * (attempt + 1))
        except Exception:
            tmp_path.unlink(missing_ok=True)
            raise

    @staticmethod
    def _next_id(items: list[dict[str, Any]]) -> int:
        if not items:
            return 1
        return max(int(item.get("id", 0)) for item in items) + 1

    # ---------- project methods ----------

    def create_project(
        self,
        project: ProjectRecord,
        include_default_tasks: bool = True,
        task_template: str = "standard",
    ) -> int:
        data = self._load_data()
        now = self._now_iso()
        new_project_id = int(data["next_project_id"])

        for existing_project in data["projects"]:
            if (
                str(existing_project.get("job_number", "")).strip() == project.job_number.strip()
                and project.job_number.strip()
            ):
                raise ValueError(
                    f"Project with job number '{project.job_number}' already exists."
                )

        project_record = {
            "id": new_project_id,
            "job_name": project.job_name.strip(),
            "job_number": project.job_number.strip(),
            "project_manager": project.project_manager.strip(),
            "sales_engineer": project.sales_engineer.strip(),
            "target_completion": self._normalize_date(project.target_completion),
            "liquid_damages": project.liquid_damages.strip(),
            "warranty_period": project.warranty_period.strip(),
            "notes": project.notes.strip(),
            "booked_date": project.booked_date.strip(),
            "group_ops_manager": project.group_ops_manager.strip(),
            "group_ops_supervisor": project.group_ops_supervisor.strip(),
            "job_subtype": project.job_subtype.strip(),
            "owner": project.owner.strip(),
            "contracted_with": project.contracted_with.strip(),
            "general_contractor": project.general_contractor.strip(),
            "contract_value": project.contract_value.strip(),
            "job_docs": project.job_docs.strip(),
            "div25_url": project.div25_url.strip(),
            "webpro_id": project.webpro_id.strip(),
            "is_test": project.is_test,
            "pinned": project.pinned,
            "rss_files": list(project.rss_files),
            "created_at": now,
            "updated_at": now,
            "created_by": self.current_user,
            "updated_by": self.current_user,
        }
        data["projects"].append(project_record)
        data["next_project_id"] = new_project_id + 1
        self._log_activity(data, new_project_id, "created", "project", project.job_name)

        if include_default_tasks:
            task_list = PHOENIX_TASKS if task_template == "phoenix" else DEFAULT_TASKS
            self._insert_default_tasks(data, new_project_id, task_list)

        self._save_data(data)
        return new_project_id

    def update_project(self, project_id: int, **changes: Any) -> None:
        allowed_fields = {
            "job_name",
            "job_number",
            "project_manager",
            "sales_engineer",
            "target_completion",
            "liquid_damages",
            "warranty_period",
            "notes",
            "booked_date",
            "group_ops_manager",
            "group_ops_supervisor",
            "job_subtype",
            "owner",
            "contracted_with",
            "general_contractor",
            "contract_value",
            "job_docs",
            "div25_url",
            "webpro_id",
            "pinned",
            "rss_files",
        }
        unknown_keys = set(changes) - allowed_fields
        if unknown_keys:
            logger.warning("update_project received unknown fields (ignored): %s", unknown_keys)

        updates = {key: value for key, value in changes.items() if key in allowed_fields}
        if not updates:
            return

        if "target_completion" in updates:
            updates["target_completion"] = self._normalize_date(updates["target_completion"])

        data = self._load_data()
        target_project = self._find_project_dict(data, project_id)
        if target_project is None:
            return

        new_job_number = str(updates.get("job_number", target_project["job_number"])).strip()
        if new_job_number:
            for existing_project in data["projects"]:
                if (
                    int(existing_project["id"]) != project_id
                    and str(existing_project.get("job_number", "")).strip() == new_job_number
                ):
                    raise ValueError(
                        f"Project with job number '{new_job_number}' already exists."
                    )

        old_values = {f: target_project.get(f) for f in updates}
        for field_name, field_value in updates.items():
            target_project[field_name] = field_value.strip() if isinstance(field_value, str) else field_value

        target_project["updated_at"] = self._now_iso()
        target_project["updated_by"] = self.current_user

        changed_fields = [
            f for f in updates
            if str(old_values[f] or "") != str(
                (updates[f].strip() if isinstance(updates[f], str) else updates[f]) or ""
            )
        ]
        if changed_fields:
            self._log_activity(
                data, project_id, "updated", "project",
                target_project["job_name"], ", ".join(changed_fields),
            )

        self._save_data(data)

    def delete_project(self, project_id: int) -> None:
        data = self._load_data()
        target = self._find_project_dict(data, project_id)
        project_name = target["job_name"] if target else f"#{project_id}"
        self._log_activity(data, project_id, "deleted", "project", project_name)
        data["projects"] = [item for item in data["projects"] if int(item["id"]) != project_id]
        data["tasks"] = [item for item in data["tasks"] if int(item["project_id"]) != project_id]
        data["activity_log"] = [a for a in data.get("activity_log", []) if int(a["project_id"]) != project_id]
        self._save_data(data)

    def get_project(self, project_id: int) -> Optional[ProjectRecord]:
        data = self._load_data()
        target_project = self._find_project_dict(data, project_id)
        return self._project_from_dict(target_project) if target_project else None

    def list_projects(
        self,
        search_text: str = "",
        include_test: bool = True,
        sort_by: str = "updated",
        sort_asc: bool = False,
    ) -> list[ProjectRecord]:
        data = self._load_data()
        search_value = search_text.strip().casefold()

        project_dicts = data["projects"]
        if not include_test:
            project_dicts = [p for p in project_dicts if not p.get("is_test", False)]
        if search_value:
            project_dicts = [
                item
                for item in project_dicts
                if search_value in str(item.get("job_name", "")).casefold()
                or search_value in str(item.get("job_number", "")).casefold()
                or search_value in str(item.get("project_manager", "")).casefold()
                or search_value in str(item.get("sales_engineer", "")).casefold()
            ]

        key_fn: Any
        if sort_by == "name":
            key_fn = lambda item: str(item.get("job_name", "")).casefold()
        elif sort_by == "job_number":
            key_fn = lambda item: str(item.get("job_number", "")).casefold()
        else:  # "updated" (default)
            key_fn = lambda item: (
                str(item.get("updated_at", "")),
                str(item.get("job_name", "")).casefold(),
            )

        project_dicts = sorted(project_dicts, key=key_fn, reverse=not sort_asc)
        # Pinned projects always float to the top, preserving sub-sort within each group
        project_dicts = sorted(project_dicts, key=lambda item: 0 if item.get("pinned") else 1)
        return [self._project_from_dict(item) for item in project_dicts]

    # ---------- task methods ----------

    def add_task(
        self,
        project_id: int,
        task_name: str,
        phase: str = "General",
        sort_order: Optional[int] = None,
        due_date: Optional[str] = None,
        completed_date: Optional[str] = None,
        notes: str = "",
    ) -> int:
        cleaned_task_name = self._clean_text(task_name)
        cleaned_phase = self._clean_text(phase) or "General"

        data = self._load_data()
        target_project = self._find_project_dict(data, project_id)
        if target_project is None:
            raise ValueError(f"Project {project_id} was not found.")

        for existing_task in data["tasks"]:
            if (
                int(existing_task["project_id"]) == project_id
                and str(existing_task["task_name"]).casefold() == cleaned_task_name.casefold()
            ):
                raise ValueError(
                    f"Task '{cleaned_task_name}' already exists for this project."
                )

        new_sort_order = (
            sort_order if sort_order is not None
            else self._next_sort_order_from_data(data, project_id)
        )
        normalized_due_date = self._normalize_date(due_date)
        normalized_completed_date = self._normalize_date(completed_date)
        new_task_id = int(data["next_task_id"])

        task_record = {
            "id": new_task_id,
            "project_id": project_id,
            "task_name": cleaned_task_name,
            "phase": cleaned_phase,
            "sort_order": int(new_sort_order),
            "due_date": normalized_due_date,
            "completed_date": normalized_completed_date,
            "is_complete": bool(normalized_completed_date),
            "notes": notes.strip(),
            "updated_by": self.current_user,
        }
        data["tasks"].append(task_record)
        data["next_task_id"] = new_task_id + 1
        target_project["updated_at"] = self._now_iso()
        target_project["updated_by"] = self.current_user
        self._log_activity(data, project_id, "created", "task", cleaned_task_name, f"phase: {cleaned_phase}")
        self._save_data(data)
        return new_task_id

    def update_task(self, task_id: int, **changes: Any) -> None:
        allowed_fields = {
            "task_name",
            "phase",
            "sort_order",
            "due_date",
            "completed_date",
            "is_complete",
            "notes",
        }
        unknown_keys = set(changes) - allowed_fields
        if unknown_keys:
            logger.warning("update_task received unknown fields (ignored): %s", unknown_keys)

        updates = {key: value for key, value in changes.items() if key in allowed_fields}
        if not updates:
            return

        if "due_date" in updates:
            updates["due_date"] = self._normalize_date(updates["due_date"])
        if "completed_date" in updates:
            updates["completed_date"] = self._normalize_date(updates["completed_date"])
        if "completed_date" in updates and "is_complete" not in updates:
            updates["is_complete"] = bool(updates["completed_date"])
        if "task_name" in updates:
            updates["task_name"] = self._clean_text(str(updates["task_name"]))
        if "phase" in updates:
            updates["phase"] = self._clean_text(str(updates["phase"])) or "General"

        data = self._load_data()
        target_task = self._find_task_dict(data, task_id)
        if target_task is None:
            return

        updated_task_name = str(updates.get("task_name", target_task["task_name"])).strip()
        for existing_task in data["tasks"]:
            if (
                int(existing_task["id"]) != task_id
                and int(existing_task["project_id"]) == int(target_task["project_id"])
                and str(existing_task["task_name"]).casefold() == updated_task_name.casefold()
            ):
                raise ValueError(
                    f"Task '{updated_task_name}' already exists for this project."
                )

        # Capture state before applying updates (needed for completion logging)
        was_complete = bool(target_task.get("is_complete"))

        for field_name, field_value in updates.items():
            target_task[field_name] = field_value

        target_task["updated_by"] = self.current_user
        owning_project = self._find_project_dict(data, int(target_task["project_id"]))
        if owning_project is not None:
            owning_project["updated_at"] = self._now_iso()
            owning_project["updated_by"] = self.current_user

        if "is_complete" in updates:
            new_complete = bool(updates["is_complete"])
            if new_complete != was_complete:
                action = "completed" if new_complete else "uncompleted"
                self._log_activity(
                    data, int(target_task["project_id"]), action, "task", target_task["task_name"]
                )
        else:
            meaningful = [k for k in updates if k != "sort_order"]
            if meaningful:
                self._log_activity(
                    data, int(target_task["project_id"]), "updated", "task",
                    target_task["task_name"], ", ".join(meaningful),
                )

        self._save_data(data)

    def delete_task(self, task_id: int) -> None:
        data = self._load_data()
        target_task = self._find_task_dict(data, task_id)
        if target_task is None:
            return

        task_name = target_task["task_name"]
        owning_project_id = int(target_task["project_id"])
        data["tasks"] = [item for item in data["tasks"] if int(item["id"]) != task_id]

        owning_project = self._find_project_dict(data, owning_project_id)
        if owning_project is not None:
            owning_project["updated_at"] = self._now_iso()
            owning_project["updated_by"] = self.current_user

        self._log_activity(data, owning_project_id, "deleted", "task", task_name)
        self._save_data(data)

    def replace_project_tasks(self, project_id: int, task_template: str) -> None:
        """Delete all tasks for a project and re-insert from the chosen template."""
        task_list = PHOENIX_TASKS if task_template == "phoenix" else DEFAULT_TASKS
        data = self._load_data()
        data["tasks"] = [t for t in data["tasks"] if int(t["project_id"]) != project_id]
        project = self._find_project_dict(data, project_id)
        if project is not None:
            project["updated_at"] = self._now_iso()
        self._insert_default_tasks(data, project_id, task_list)
        self._save_data(data)

    def list_tasks(self, project_id: int, phase: Optional[str] = None) -> list[TaskRecord]:
        data = self._load_data()
        task_dicts = [item for item in data["tasks"] if int(item["project_id"]) == project_id]

        if phase:
            task_dicts = [item for item in task_dicts if str(item["phase"]) == phase]

        task_dicts = sorted(
            task_dicts,
            key=lambda item: (
                int(item.get("sort_order", 0)),
                str(item.get("task_name", "")).casefold(),
            ),
        )
        return [self._task_from_dict(item) for item in task_dicts]

    def set_task_completed(
        self,
        task_id: int,
        completed: bool,
        completed_date: Optional[str] = None,
    ) -> None:
        if completed and not completed_date:
            completed_date = date.today().isoformat()
        self.update_task(
            task_id,
            is_complete=bool(completed),
            completed_date=completed_date if completed else None,
        )

    # ---------- note methods ----------

    def add_note(self, project_id: int, content: str, note_date: str = "",
                 status: str = "Open", closeout_comment: str = "") -> int:
        data = self._load_data()
        if self._find_project_dict(data, project_id) is None:
            raise ValueError(f"Project {project_id} not found.")
        new_id = int(data.get("next_note_id", 1))
        existing = [n for n in data.get("notes", []) if int(n["project_id"]) == project_id]
        note_number = len(existing) + 1
        data.setdefault("notes", []).append({
            "id": new_id, "project_id": project_id,
            "note_number": note_number, "date": note_date.strip(),
            "content": content.strip(), "status": status,
            "closeout_comment": closeout_comment.strip(),
        })
        data["next_note_id"] = new_id + 1
        self._save_data(data)
        return new_id

    def update_note(self, note_id: int, **changes: Any) -> None:
        allowed = {"date", "content", "status", "closeout_comment"}
        updates = {k: v for k, v in changes.items() if k in allowed}
        if not updates:
            return
        data = self._load_data()
        note = next((n for n in data.get("notes", []) if int(n["id"]) == note_id), None)
        if note is None:
            return
        for k, v in updates.items():
            note[k] = v.strip() if isinstance(v, str) else v
        self._save_data(data)

    def delete_note(self, note_id: int) -> None:
        data = self._load_data()
        data["notes"] = [n for n in data.get("notes", []) if int(n["id"]) != note_id]
        # Re-number remaining notes per project
        by_project: dict[int, list] = {}
        for n in data["notes"]:
            by_project.setdefault(int(n["project_id"]), []).append(n)
        for notes_list in by_project.values():
            for i, n in enumerate(notes_list, start=1):
                n["note_number"] = i
        self._save_data(data)

    def list_notes(self, project_id: int) -> list[NoteRecord]:
        data = self._load_data()
        return [
            NoteRecord(
                id=n["id"], project_id=n["project_id"],
                note_number=n["note_number"], date=n["date"],
                content=n["content"], status=n["status"],
                closeout_comment=n.get("closeout_comment", ""),
            )
            for n in sorted(data.get("notes", []),
                            key=lambda x: int(x["note_number"]))
            if int(n["project_id"]) == project_id
        ]

    # ---------- change order methods ----------

    @staticmethod
    def _co_to_dict(co: ChangeOrderRecord, project_id: int, new_id: int) -> dict[str, Any]:
        return {
            "id": new_id, "project_id": project_id,
            "cop_number": co.cop_number.strip(),
            "reference": co.reference.strip(),
            "description": co.description.strip(),
            "creation_date": co.creation_date.strip(),
            "ats_price": co.ats_price.strip(),
            "ats_direct_cost": co.ats_direct_cost.strip(),
            "ats_status": co.ats_status,
            "booked_in_portal": co.booked_in_portal.strip(),
            "ats_booked_co": co.ats_booked_co.strip(),
            "mech_co": co.mech_co.strip(),
            "sub_quoted_price": co.sub_quoted_price.strip(),
            "sub_plug_number": co.sub_plug_number.strip(),
            "sub_status": co.sub_status,
            "sub_co_sent": co.sub_co_sent.strip(),
            "sub_co_number": co.sub_co_number.strip(),
            "notes": co.notes.strip(),
        }

    @staticmethod
    def _co_from_dict(d: dict[str, Any]) -> ChangeOrderRecord:
        return ChangeOrderRecord(
            id=d["id"], project_id=d["project_id"],
            cop_number=d.get("cop_number", ""),
            reference=d.get("reference", ""),
            description=d.get("description", ""),
            creation_date=d.get("creation_date", ""),
            ats_price=d.get("ats_price", ""),
            ats_direct_cost=d.get("ats_direct_cost", ""),
            ats_status=d.get("ats_status", "Pending"),
            booked_in_portal=d.get("booked_in_portal", ""),
            ats_booked_co=d.get("ats_booked_co", ""),
            mech_co=d.get("mech_co", ""),
            sub_quoted_price=d.get("sub_quoted_price", ""),
            sub_plug_number=d.get("sub_plug_number", ""),
            sub_status=d.get("sub_status", "Pending"),
            sub_co_sent=d.get("sub_co_sent", ""),
            sub_co_number=d.get("sub_co_number", ""),
            notes=d.get("notes", ""),
        )

    def add_change_order(self, project_id: int, co: ChangeOrderRecord) -> int:
        data = self._load_data()
        if self._find_project_dict(data, project_id) is None:
            raise ValueError(f"Project {project_id} not found.")
        new_id = int(data.get("next_co_id", 1))
        data.setdefault("change_orders", []).append(
            self._co_to_dict(co, project_id, new_id)
        )
        data["next_co_id"] = new_id + 1
        self._save_data(data)
        return new_id

    def update_change_order(self, co_id: int, co: ChangeOrderRecord) -> None:
        data = self._load_data()
        rec = next((c for c in data.get("change_orders", []) if int(c["id"]) == co_id), None)
        if rec is None:
            return
        updated = self._co_to_dict(co, int(rec["project_id"]), co_id)
        rec.update(updated)
        self._save_data(data)

    def delete_change_order(self, co_id: int) -> None:
        data = self._load_data()
        data["change_orders"] = [c for c in data.get("change_orders", [])
                                  if int(c["id"]) != co_id]
        self._save_data(data)

    def list_change_orders(self, project_id: int) -> list[ChangeOrderRecord]:
        data = self._load_data()
        return [
            self._co_from_dict(c)
            for c in data.get("change_orders", [])
            if int(c["project_id"]) == project_id
        ]

    def get_co_summary(self, project_id: int) -> dict[str, Any]:
        """Return ATS and Sub contract totals for the summary bar."""
        cos = self.list_change_orders(project_id)
        project = self.get_project(project_id)

        def _parse(val: str) -> float:
            try:
                return float(str(val).replace(",", "").replace("$", "").strip())
            except (ValueError, TypeError):
                return 0.0

        ats_base = _parse(project.contract_value) if project else 0.0
        ats_accepted = sum(_parse(c.ats_price) for c in cos if c.ats_status == "Accepted")
        ats_pending  = sum(_parse(c.ats_price) for c in cos if c.ats_status == "Pending")

        sub_base     = 0.0  # stored in project if needed
        sub_accepted = sum(
            _parse(c.sub_quoted_price) if c.sub_quoted_price else _parse(c.sub_plug_number)
            for c in cos if c.sub_status == "Accepted"
        )
        sub_pending  = sum(
            _parse(c.sub_quoted_price) if c.sub_quoted_price else _parse(c.sub_plug_number)
            for c in cos if c.sub_status == "Pending"
        )

        return {
            "ats_base":         ats_base,
            "ats_accepted":     ats_accepted,
            "ats_pending":      ats_pending,
            "ats_current":      ats_base + ats_accepted,
            "sub_base":         sub_base,
            "sub_accepted":     sub_accepted,
            "sub_pending":      sub_pending,
            "sub_current":      sub_base + sub_accepted,
        }

    def get_project_summary(self, project_id: int) -> dict[str, Any]:
        project_record = self.get_project(project_id)
        task_records = self.list_tasks(project_id)
        total_tasks = len(task_records)
        completed_tasks = sum(1 for item in task_records if item.is_complete)
        pending_tasks = total_tasks - completed_tasks
        progress_percent = (
            round((completed_tasks / total_tasks) * 100, 1) if total_tasks else 0.0
        )

        phase_breakdown: dict[str, dict[str, int]] = {}
        for task_record in task_records:
            bucket = phase_breakdown.setdefault(
                task_record.phase, {"total": 0, "completed": 0, "pending": 0}
            )
            bucket["total"] += 1
            if task_record.is_complete:
                bucket["completed"] += 1
            else:
                bucket["pending"] += 1

        return {
            "project": asdict(project_record) if project_record else None,
            "totals": {
                "tasks": total_tasks,
                "completed": completed_tasks,
                "pending": pending_tasks,
                "progress_percent": progress_percent,
            },
            "phase_breakdown": phase_breakdown,
        }

    # ---------- workbook import ----------

    def import_project_from_workbook(
        self,
        workbook_path: str | Path,
        sheet_name: Optional[str] = None,
        create_missing_tasks: bool = True,
    ) -> int:
        workbook_file = Path(workbook_path)
        workbook = load_workbook(workbook_file, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

        job_number = self._value(sheet, "H3")
        if not job_number:
            logger.warning(
                "Imported workbook '%s' has no job number in cell H3. "
                "Duplicate detection will be skipped for this import.",
                workbook_file.name,
            )

        project = ProjectRecord(
            job_name=self._value(sheet, "C3"),
            job_number=job_number,
            project_manager=self._value(sheet, "C4"),
            sales_engineer=self._value(sheet, "H4"),
            target_completion=self._value(sheet, "E5"),
            liquid_damages=self._value(sheet, "E6"),
            warranty_period=self._value(sheet, "E7"),
            notes=f"Imported from workbook: {workbook_file.name}",
        )

        imported_project_id = self.create_project(project, include_default_tasks=True)
        existing_tasks = {
            task.task_name.casefold(): task
            for task in self.list_tasks(imported_project_id)
        }

        imported_task_items = self._extract_tasks_from_sheet(sheet)
        for imported_item in imported_task_items:
            task_key = imported_item["task_name"].casefold()
            existing_task = existing_tasks.get(task_key)
            if existing_task and existing_task.id is not None:
                self.update_task(
                    existing_task.id,
                    phase=imported_item["phase"],
                    sort_order=imported_item["sort_order"],
                    completed_date=imported_item["completed_date"],
                    is_complete=bool(imported_item["completed_date"]),
                )
            elif create_missing_tasks:
                self.add_task(
                    project_id=imported_project_id,
                    task_name=imported_item["task_name"],
                    phase=imported_item["phase"],
                    sort_order=imported_item["sort_order"],
                    completed_date=imported_item["completed_date"],
                )

        return imported_project_id

    def create_test_jobs(self) -> None:
        """Create 5 demo jobs tagged is_test=True for training purposes."""

        def _make(job_name: str, job_number: str, **kwargs: Any) -> int:
            template = kwargs.pop("task_template", "standard")
            rec = ProjectRecord(job_name=job_name, job_number=job_number, is_test=True, **kwargs)
            return self.create_project(rec, include_default_tasks=True, task_template=template)

        # ── Job 1: Early-stage standard job ──────────────────────────────────
        pid1 = _make(
            "PNNL - Building 3000 Controls Upgrade",
            "DEMO-1001",
            project_manager="Justin Glave",
            sales_engineer="Sarah Mitchell",
            booked_date="2026-01-10",
            contract_value="248500",
            owner="Pacific Northwest National Laboratory",
            contracted_with="ATS Inc.",
            general_contractor="Hensel Phelps",
            warranty_period="12 months",
            liquid_damages="$500/day",
            target_completion="2026-09-30",
            notes="Kickoff meeting scheduled for Feb 3. Owner rep is Dan Schultz.",
        )
        # Mark a few early tasks complete
        tasks1 = self.list_tasks(pid1)
        for t in tasks1[:4]:
            self.set_task_completed(t.id, True)  # type: ignore[arg-type]
        # Notes
        self.add_note(pid1, "Kickoff meeting held Feb 3. All stakeholders present. Schedule confirmed.", "2026-02-03")
        self.add_note(pid1, "Owner requested 2-week schedule acceleration. PM reviewing impacts.", "2026-02-18")
        # Change order
        self.add_change_order(pid1, ChangeOrderRecord(
            cop_number="CO-001", description="Added 6 additional DDC points per owner request",
            ats_price="8400", ats_status="Accepted", creation_date="2026-02-20",
        ))

        # ── Job 2: Mid-progress Phoenix job ──────────────────────────────────
        pid2 = _make(
            "Hanford Site - HVAC Controls Replacement",
            "DEMO-1002",
            project_manager="Lisa Park",
            sales_engineer="Tom Nguyen",
            booked_date="2025-11-05",
            contract_value="512000",
            owner="US Department of Energy",
            contracted_with="ATS Inc.",
            general_contractor="Bechtel",
            warranty_period="24 months",
            liquid_damages="$1,000/day",
            target_completion="2026-07-15",
            job_subtype="Phoenix",
            task_template="phoenix",
            notes="Security clearance required for all on-site personnel.",
        )
        tasks2 = self.list_tasks(pid2)
        for t in tasks2[:10]:
            self.set_task_completed(t.id, True)  # type: ignore[arg-type]
        self.add_note(pid2, "Security badges issued to all ATS personnel.", "2025-11-20")
        self.add_note(pid2, "Materials delivered to laydown area. Valves inspected — 2 units damaged in transit, replacements ordered.", "2026-01-08")
        self.add_note(pid2, "Electrical sub mobilized. Install began in mechanical room B.", "2026-02-01")
        self.add_change_order(pid2, ChangeOrderRecord(
            cop_number="CO-001", description="Replacement of 2 damaged valves damaged in shipping",
            ats_price="3200", ats_status="Accepted", creation_date="2026-01-10",
        ))
        self.add_change_order(pid2, ChangeOrderRecord(
            cop_number="CO-002", description="Extended rigging for rooftop AHU installation",
            ats_price="11750", ats_status="Pending", creation_date="2026-02-05",
        ))

        # ── Job 3: Nearly complete standard job ──────────────────────────────
        pid3 = _make(
            "Boeing Renton - Building 4-20 BAS",
            "DEMO-1003",
            project_manager="Justin Glave",
            sales_engineer="Karen Ortiz",
            booked_date="2025-08-12",
            contract_value="189000",
            owner="Boeing Commercial Airplanes",
            contracted_with="ATS Inc.",
            general_contractor="Turner Construction",
            warranty_period="12 months",
            target_completion="2026-04-01",
            notes="All punch list items resolved. Awaiting final owner sign-off.",
        )
        tasks3 = self.list_tasks(pid3)
        for t in tasks3[:-4]:
            self.set_task_completed(t.id, True)  # type: ignore[arg-type]
        self.add_note(pid3, "Commissioning complete. All sequences verified by owner.", "2026-02-28")
        self.add_note(pid3, "Final punch list walkthrough completed. 3 minor items remain.", "2026-03-10")
        self.add_note(pid3, "All punch items resolved. Owner training delivered to 4 staff members.", "2026-03-18")
        self.add_change_order(pid3, ChangeOrderRecord(
            cop_number="CO-001", description="Graphic screen additions — 4 custom floor plan views",
            ats_price="5600", ats_status="Accepted", creation_date="2025-10-15",
        ))
        self.add_change_order(pid3, ChangeOrderRecord(
            cop_number="CO-002", description="Added trend logging for 12 additional points",
            ats_price="2100", ats_status="Accepted", creation_date="2025-12-01",
        ))
        self.add_change_order(pid3, ChangeOrderRecord(
            cop_number="CO-003", description="Extended warranty to 24 months per owner request",
            ats_price="4500", ats_status="Accepted", creation_date="2026-01-20",
        ))

        # ── Job 4: Brand new job, just kicked off ────────────────────────────
        pid4 = _make(
            "Microsoft Campus - Lab 7 Automation",
            "DEMO-1004",
            project_manager="Rachel Simmons",
            sales_engineer="Justin Glave",
            booked_date="2026-03-01",
            contract_value="375000",
            owner="Microsoft Corporation",
            contracted_with="ATS Inc.",
            warranty_period="12 months",
            liquid_damages="$750/day",
            target_completion="2026-12-15",
            notes="New customer relationship. High visibility project — weekly status reports required.",
        )
        tasks4 = self.list_tasks(pid4)
        for t in tasks4[:2]:
            self.set_task_completed(t.id, True)  # type: ignore[arg-type]
        self.add_note(pid4, "Sales-Ops turnover meeting held. PM introduced to owner contact.", "2026-03-05")
        self.add_change_order(pid4, ChangeOrderRecord(
            cop_number="CO-001", description="Scope addition: integrate lab fume hood monitoring",
            ats_price="22000", ats_status="Pending", creation_date="2026-03-15",
        ))

        # ── Job 5: Fully closed out Phoenix job ──────────────────────────────
        pid5 = _make(
            "Richland Schools - District HVAC Controls",
            "DEMO-1005",
            project_manager="Lisa Park",
            sales_engineer="Tom Nguyen",
            booked_date="2025-03-20",
            contract_value="298000",
            owner="Richland School District",
            contracted_with="ATS Inc.",
            general_contractor="Lydig Construction",
            warranty_period="24 months",
            target_completion="2025-12-01",
            job_subtype="Phoenix",
            task_template="phoenix",
            notes="Project closed. Warranty period active through Dec 2027.",
        )
        tasks5 = self.list_tasks(pid5)
        for t in tasks5:
            self.set_task_completed(t.id, True)  # type: ignore[arg-type]
        self.add_note(pid5, "Project kick-off April 2025. Owner very engaged.", "2025-04-02")
        self.add_note(pid5, "All 8 schools commissioned and signed off.", "2025-11-14")
        self.add_note(pid5, "Final billing submitted. Retainage released.", "2025-12-10")
        self.add_change_order(pid5, ChangeOrderRecord(
            cop_number="CO-001", description="Added controls for 3 portable classrooms",
            ats_price="9800", ats_status="Accepted", creation_date="2025-06-10",
        ))
        self.add_change_order(pid5, ChangeOrderRecord(
            cop_number="CO-002", description="Overtime labor — accelerated schedule for winter break deadline",
            ats_price="6300", ats_status="Accepted", creation_date="2025-10-01",
        ))

    def delete_test_jobs(self) -> None:
        """Remove all projects tagged is_test=True and their associated data."""
        data = self._load_data()
        test_ids = {int(p["id"]) for p in data["projects"] if p.get("is_test", False)}
        if not test_ids:
            return
        data["projects"] = [p for p in data["projects"] if int(p["id"]) not in test_ids]
        data["tasks"] = [t for t in data["tasks"] if int(t["project_id"]) not in test_ids]
        data["notes"] = [n for n in data["notes"] if int(n["project_id"]) not in test_ids]
        data["change_orders"] = [c for c in data["change_orders"] if int(c["project_id"]) not in test_ids]
        data["activity_log"] = [a for a in data.get("activity_log", []) if int(a["project_id"]) not in test_ids]
        self._save_data(data)

    def export_project_snapshot(self, project_id: int, export_path: str | Path) -> Path:
        output_file = Path(export_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "summary": self.get_project_summary(project_id),
            "tasks": [asdict(task) for task in self.list_tasks(project_id)],
        }
        output_file.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        return output_file

    def export_project_to_excel(self, project_id: int, export_path: str | Path) -> Path:
        """Export project info and all tasks to a formatted Excel workbook."""
        from openpyxl import Workbook
        wb = Workbook()
        self._write_project_sheets(wb, project_id)
        output_file = Path(export_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        return output_file

    def export_projects_to_excel(self, project_ids: list[int], export_path: str | Path) -> Path:
        """Export multiple projects to a single workbook, one set of sheets per project."""
        from openpyxl import Workbook
        wb = Workbook()
        # Remove the default empty sheet
        wb.remove(wb.active)  # type: ignore[arg-type]
        for pid in project_ids:
            project = self.get_project(pid)
            if project is None:
                continue
            prefix = project.job_number or project.job_name
            self._write_project_sheets(wb, pid, prefix=prefix)
        output_file = Path(export_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        return output_file

    def _write_project_sheets(self, wb: Any, project_id: int, prefix: str = "") -> None:
        """Write project report sheets into wb. prefix="" uses standard sheet names;
        a non-empty prefix creates uniquely-named sheets for multi-project workbooks."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        project = self.get_project(project_id)
        if not project:
            raise ValueError(f"Project {project_id} not found.")
        tasks   = self.list_tasks(project_id)
        summary = self.get_project_summary(project_id)
        totals  = summary["totals"]

        # Sheet name helpers
        def _sheet_name(suffix: str) -> str:
            if not prefix:
                return suffix
            safe = "".join(c for c in prefix if c not in r'\/*?:[]')
            return f"{safe[:24]} {suffix}"[:31]

        if not prefix:
            ws = wb.active
            ws.title = "Project Report"
        else:
            ws = wb.create_sheet(_sheet_name("Tasks"))

        dark_blue      = "1E3A5F"
        mid_blue       = "2D5A8E"
        light_blue     = "D6E4F0"
        complete_green = "E8F5E9"
        white          = "FFFFFF"
        gray           = "F2F2F2"

        def hfont(size: int = 11, color: str = white) -> Font:
            return Font(name="Segoe UI", size=size, bold=True, color=color)

        def cfont(bold: bool = False, color: str = "000000") -> Font:
            return Font(name="Segoe UI", size=10, bold=bold, color=color)

        def solid(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color)

        def thin_border() -> Border:
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        def _sc(cell: Any, *, fnt: Any = None, fil: Any = None,
                aln: Any = None, brd: Any = None) -> None:
            """Style a cell — type: ignore suppresses MergedCell union warnings."""
            if fnt is not None: cell.font      = fnt  # type: ignore[union-attr]
            if fil is not None: cell.fill      = fil  # type: ignore[union-attr]
            if aln is not None: cell.alignment = aln  # type: ignore[union-attr]
            if brd is not None: cell.border    = brd  # type: ignore[union-attr]

        def _sec(merge_range: str, title_text: str, row_num: int) -> None:
            """Write a formatted section header row."""
            ws.merge_cells(merge_range)
            ws[f"A{row_num}"] = title_text  # type: ignore[index]
            _sc(ws[f"A{row_num}"], fnt=hfont(10), fil=solid(mid_blue), aln=centre)  # type: ignore[index]
            ws.row_dimensions[row_num].height = 18

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Project Report — {project.job_name}"  # type: ignore[index]
        _sc(ws["A1"], fnt=hfont(14), fil=solid(dark_blue), aln=centre)  # type: ignore[index]
        ws.row_dimensions[1].height = 28

        # Project info
        info_fields = [
            ("Job Name",             project.job_name),
            ("Job Number",           project.job_number),
            ("Project Manager",      project.project_manager),
            ("Sales Engineer",       project.sales_engineer),
            ("Target Completion",    project.target_completion or "—"),
            ("Booked Date",          project.booked_date or "—"),
            ("Contract Value",       f"${float(project.contract_value):,.2f}"
                                     if project.contract_value else "—"),
            ("Owner",                project.owner or "—"),
            ("Contracted With",      project.contracted_with or "—"),
            ("General Contractor",   project.general_contractor or "—"),
            ("Group Ops Manager",    project.group_ops_manager or "—"),
            ("Group Ops Supervisor", project.group_ops_supervisor or "—"),
            ("Job Sub-Type",         project.job_subtype or "—"),
            ("Warranty Period",      project.warranty_period or "—"),
            ("Liquid Damages",       project.liquid_damages or "—"),
            ("Div25 URL",            project.div25_url or "—"),
            ("Job Docs",             project.job_docs or "—"),
            ("Notes",                project.notes or "—"),
        ]

        row = 2
        _sec(f"A{row}:H{row}", "PROJECT INFORMATION", row)
        row += 1

        for i, (lbl, val) in enumerate(info_fields):
            off = 0 if i % 2 == 0 else 4
            r   = row + (i // 2)
            lc  = ws.cell(row=r, column=1 + off, value=lbl)
            vc  = ws.cell(row=r, column=2 + off, value=val)
            _sc(lc, fnt=cfont(bold=True), fil=solid(light_blue), aln=left, brd=thin_border())
            _sc(vc, fnt=cfont(), fil=solid(gray if i % 2 == 0 else white), aln=left, brd=thin_border())
            ws.merge_cells(f"{vc.column_letter}{r}:{chr(ord(vc.column_letter)+2)}{r}")  # type: ignore[union-attr]

        row += (len(info_fields) + 1) // 2 + 1

        # Summary
        _sec(f"A{row}:H{row}", "TASK SUMMARY", row)
        row += 1
        for i, (sl, sv) in enumerate([
            ("Total Tasks", totals["tasks"]),
            ("Completed",   totals["completed"]),
            ("Pending",     totals["pending"]),
            ("Progress",    f"{totals['progress_percent']}%"),
        ]):
            lc = ws.cell(row=row, column=1 + i * 2, value=sl)
            vc = ws.cell(row=row, column=2 + i * 2, value=sv)
            _sc(lc, fnt=cfont(bold=True), fil=solid(light_blue), aln=centre, brd=thin_border())
            _sc(vc, fnt=cfont(bold=True), fil=solid(gray),        aln=centre, brd=thin_border())
        row += 2

        # Tasks
        _sec(f"A{row}:H{row}", "TASKS", row)
        row += 1
        for ci, (txt, cw) in enumerate(zip(
            ["#", "Task Name", "Phase", "Status", "Completed Date", "Notes"],
            [5, 45, 18, 12, 16, 40],
        ), start=1):
            c = ws.cell(row=row, column=ci, value=txt)
            _sc(c, fnt=hfont(10), fil=solid(dark_blue), aln=centre, brd=thin_border())
            ws.column_dimensions[c.column_letter].width = cw  # type: ignore[union-attr]
        row += 1

        for idx, task in enumerate(tasks, start=1):
            rf = solid(complete_green) if task.is_complete else solid(white)
            for ci, v in enumerate(
                [idx, task.task_name, task.phase,
                 "\u2713 Complete" if task.is_complete else "Pending",
                 task.completed_date or "", task.notes or ""],
                start=1,
            ):
                c = ws.cell(row=row, column=ci, value=v)
                _sc(c, fnt=cfont(color="1E4D2B" if task.is_complete else "000000"),
                    fil=rf, aln=left, brd=thin_border())
            row += 1

        ws.freeze_panes = f"A{row - len(tasks)}"

        # ── Notes tab ─────────────────────────────────────────────────────────
        notes = self.list_notes(project_id)
        wn = wb.create_sheet(_sheet_name("Job Progress Notes") if prefix else "Job Progress Notes")

        # Header block
        wn.merge_cells("B1:F1")
        wn["B1"] = "Job Progress Notes"  # type: ignore[index]
        _sc(wn["B1"], fnt=Font(name="Segoe UI", size=13, bold=True, color="000000"),  # type: ignore[index]
            fil=solid("FFFF00"), aln=left)
        wn.row_dimensions[1].height = 20

        for r, (lbl, val) in enumerate([
            ("Project Name:", project.job_name),
            ("Project Number:", project.job_number),
            ("PM:", project.project_manager),
        ], start=2):
            lc = wn.cell(row=r, column=3, value=lbl)
            vc = wn.cell(row=r, column=4, value=val)
            _sc(lc, fnt=Font(name="Segoe UI", size=10, bold=True, color="000000"),
                fil=solid("FFFF00"), aln=left)
            _sc(vc, fnt=Font(name="Segoe UI", size=10, color="000000"), aln=left)
            wn.row_dimensions[r].height = 16

        # Column headers
        note_hdrs = ["Note", "Date", "Note or Comment", "", "", "Status", "Close out Comment"]
        note_widths = [6, 14, 30, 30, 20, 10, 50]
        for ci, (h, cw) in enumerate(zip(note_hdrs, note_widths), start=1):
            c = wn.cell(row=5, column=ci, value=h)
            _sc(c, fnt=Font(name="Segoe UI", size=10, bold=True, color="000000"),
                fil=solid("FFFF00"), aln=centre,
                brd=thin_border())
            wn.column_dimensions[c.column_letter].width = cw  # type: ignore[union-attr]
        wn.row_dimensions[5].height = 18

        # Note rows
        for nr, note in enumerate(notes, start=6):
            is_open = note.status == "Open"
            row_bg  = "FFCCCC" if is_open else "CCFFCC"
            status_color = "FF0000" if is_open else "006400"
            row_vals: list[Any] = [
                note.note_number, note.date, note.content,
                "", "", note.status, note.closeout_comment,
            ]
            for ci, v in enumerate(row_vals, start=1):
                c = wn.cell(row=nr, column=ci, value=v)
                if ci == 6:  # Status column
                    _sc(c, fnt=Font(name="Segoe UI", size=10, bold=True, color=status_color),
                        fil=solid(row_bg), aln=centre, brd=thin_border())
                else:
                    _sc(c, fnt=Font(name="Segoe UI", size=10, color="000000"),
                        fil=solid(row_bg if ci != 3 else white), aln=left, brd=thin_border())
            # Merge note content across columns C-E
            wn.merge_cells(f"C{nr}:E{nr}")
            wn.row_dimensions[nr].height = 30

        wn.freeze_panes = "A6"

        # ── Change Orders tab ─────────────────────────────────────────────────
        cos = self.list_change_orders(project_id)
        co_sum = self.get_co_summary(project_id)
        wco = wb.create_sheet(_sheet_name("ATS CO Log") if prefix else "ATS CO Log")

        # Summary block
        summary_rows = [
            ("ATS Base Contract",       f"${co_sum['ats_base']:,.2f}",
             "Sub Base Contract",        f"${co_sum['sub_base']:,.2f}"),
            ("Total ATS Accepted CO",   f"${co_sum['ats_accepted']:,.2f}",
             "Total Accepted Sub CO",   f"${co_sum['sub_accepted']:,.2f}"),
            ("ATS Pending",             f"${co_sum['ats_pending']:,.2f}",
             "Sub Pending",             f"${co_sum['sub_pending']:,.2f}"),
            ("ATS Current Contract",    f"${co_sum['ats_current']:,.2f}",
             "Sub Current Contract",    f"${co_sum['sub_current']:,.2f}"),
        ]
        for r, (l1, v1, l2, v2) in enumerate(summary_rows, start=1):
            for ci, val in enumerate([l1, v1, "", "", "", "", l2, v2], start=1):
                c = wco.cell(row=r, column=ci, value=val)
                is_label = ci in (1, 7)
                _sc(c, fnt=Font(name="Segoe UI", size=10,
                                bold=is_label, color="000000"),
                    fil=solid(light_blue if is_label else white), aln=left)

        # CO column headers — row 6
        co_headers = [
            "COP#", "Reference", "Description", "Creation Date",
            "ATS Price", "ATS Direct Cost", "ATS Status", "Booked in Portal",
            "ATS Booked CO#", "Mech CO#",
            "Sub Quoted Price", "Sub Plug #", "Sub Price (Auto)",
            "Sub Status", "Sub CO Sent", "Sub CO#", "Notes",
        ]
        co_widths = [10, 14, 40, 14, 12, 14, 14, 14, 12, 10, 14, 12, 14, 12, 12, 10, 30]
        for ci, (h, cw) in enumerate(zip(co_headers, co_widths), start=1):
            c = wco.cell(row=6, column=ci, value=h)
            _sc(c, fnt=hfont(10), fil=solid(dark_blue), aln=centre, brd=thin_border())
            wco.column_dimensions[c.column_letter].width = cw  # type: ignore[union-attr]

        # CO data rows
        for ri, co in enumerate(cos, start=7):
            sub_price = co.sub_quoted_price if co.sub_quoted_price else co.sub_plug_number
            status_colors = {"Accepted": "CCFFCC", "Rejected": "FFCCCC", "Pending": "FFFACD"}
            ats_bg = status_colors.get(co.ats_status, white)
            sub_bg = status_colors.get(co.sub_status, white)
            row_vals = [
                co.cop_number, co.reference, co.description, co.creation_date,
                co.ats_price, co.ats_direct_cost, co.ats_status, co.booked_in_portal,
                co.ats_booked_co, co.mech_co,
                co.sub_quoted_price, co.sub_plug_number, sub_price,
                co.sub_status, co.sub_co_sent, co.sub_co_number, co.notes,
            ]
            for ci, v in enumerate(row_vals, start=1):
                bg = ats_bg if ci <= 10 else (sub_bg if ci <= 16 else white)
                c = wco.cell(row=ri, column=ci, value=v)
                _sc(c, fnt=cfont(), fil=solid(bg), aln=left, brd=thin_border())

        wco.freeze_panes = "A7"

    # ---------- email import ----------

    def import_project_from_email(
        self, eml_path: str | Path
    ) -> tuple[ProjectRecord, bool]:
        """Parse an Odin assignment .eml file and return a ProjectRecord.

        Returns (record, is_duplicate) where is_duplicate=True means a project
        with that job number already exists in the database.
        """
        import email as _email
        import html as _html
        import re as _re
        from html.parser import HTMLParser as _HTMLParser

        eml_text = Path(eml_path).read_bytes()
        msg = _email.message_from_bytes(eml_text)

        # Get HTML body
        html_body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/html":
                    raw = part.get_payload(decode=True)
                    if isinstance(raw, bytes):
                        html_body = raw.decode("utf-8", errors="replace")
                    break
        else:
            payload = msg.get_payload(decode=True)
            if isinstance(payload, bytes):
                html_body = payload.decode("utf-8", errors="replace")

        if not html_body:
            raise ValueError("No HTML content found in the email.")

        # Parse key:value pairs from the HTML table
        class _TableParser(_HTMLParser):
            def __init__(self) -> None:
                super().__init__()
                self.data: dict[str, str] = {}
                self._cells: list[str] = []
                self._in_td = False
                self._current = ""

            def handle_starttag(self, tag: str, attrs: Any) -> None:
                if tag == "tr":
                    self._cells = []
                elif tag == "td":
                    self._in_td = True
                    self._current = ""

            def handle_endtag(self, tag: str) -> None:
                if tag == "td":
                    self._in_td = False
                    self._cells.append(self._current.strip())
                elif tag == "tr" and len(self._cells) == 2:
                    key = self._cells[0].strip()
                    val = self._cells[1].strip()
                    if key:
                        self.data[key] = val

            def handle_data(self, text: str) -> None:
                if self._in_td:
                    self._current += text

            def handle_entityref(self, name):
                if self._in_td:
                    self._current += _html.unescape(f"&{name};")

            def handle_charref(self, name):
                if self._in_td:
                    self._current += _html.unescape(f"&#{name};")

        # Decode quoted-printable encoding
        import quopri as _quopri
        decoded = _quopri.decodestring(html_body.encode()).decode("utf-8", errors="replace")

        parser = _TableParser()
        parser.feed(decoded)
        d = parser.data

        def get(*keys: str) -> str:
            for k in keys:
                val = d.get(k, "").strip()
                if val:
                    return val
            return ""

        # Extract booking notes — strip timestamp prefix if present
        raw_notes = get("Booking Notes")
        notes_match = _re.sub(r"^\d+/\d+/\d{4}.*?\n", "", raw_notes).strip()

        record = ProjectRecord(
            job_name       = get("Job Name"),
            job_number     = get("Job Number"),
            project_manager= get("Project Manager"),
            sales_engineer = get("Sales Person"),
            booked_date    = get("Booked"),
            group_ops_manager   = get("Group Operations Manager"),
            group_ops_supervisor= get("Group Operations Supervisor"),
            job_subtype    = get("Job Sub-Type"),
            owner          = get("Owner"),
            contracted_with= get("Contracted With"),
            general_contractor  = get("General Contractor"),
            contract_value = get("Contract Value"),
            job_docs       = get("Job Docs"),
            div25_url      = get("Div25 URL"),
            notes          = notes_match or raw_notes,
        )

        # Check for duplicate
        data = self._load_data()
        is_duplicate = any(
            str(p.get("job_number", "")).strip() == record.job_number.strip()
            for p in data["projects"]
            if record.job_number.strip()
        )

        return record, is_duplicate

    def update_project_from_email(self, project_id: int, record: ProjectRecord) -> None:
        """Update an existing project with fields parsed from an email."""
        self.update_project(
            project_id,
            job_name=record.job_name,
            job_number=record.job_number,
            project_manager=record.project_manager,
            sales_engineer=record.sales_engineer,
            booked_date=record.booked_date,
            group_ops_manager=record.group_ops_manager,
            group_ops_supervisor=record.group_ops_supervisor,
            job_subtype=record.job_subtype,
            owner=record.owner,
            contracted_with=record.contracted_with,
            general_contractor=record.general_contractor,
            contract_value=record.contract_value,
            job_docs=record.job_docs,
            div25_url=record.div25_url,
            notes=record.notes,
        )

    # ---------- activity log ----------

    def _log_activity(
        self,
        data: dict[str, Any],
        project_id: int,
        action: str,
        entity_type: str,
        entity_name: str,
        details: str = "",
    ) -> None:
        new_id = int(data.get("next_activity_id", 1))
        data.setdefault("activity_log", []).append({
            "id": new_id,
            "project_id": project_id,
            "timestamp": self._now_iso(),
            "user": self.current_user,
            "action": action,
            "entity_type": entity_type,
            "entity_name": entity_name,
            "details": details,
        })
        data["next_activity_id"] = new_id + 1

    def log_rss_proposal(self, project_id: int, table_name: str, details: str) -> None:
        data = self._load_data()
        self._log_activity(data, project_id, "proposed", "rss", table_name, details)
        self._save_data(data)

    def list_activity(self, project_id: int) -> list[ActivityRecord]:
        data = self._load_data()
        entries = [
            e for e in data.get("activity_log", [])
            if int(e["project_id"]) == project_id
        ]
        entries.sort(key=lambda e: e.get("timestamp", ""), reverse=True)
        return [
            ActivityRecord(
                id=e["id"],
                project_id=e["project_id"],
                timestamp=e["timestamp"],
                user=e["user"],
                action=e["action"],
                entity_type=e["entity_type"],
                entity_name=e["entity_name"],
                details=e.get("details", ""),
            )
            for e in entries
        ]

    def delete_activity(self, activity_id: int) -> None:
        data = self._load_data()
        data["activity_log"] = [
            a for a in data.get("activity_log", [])
            if int(a["id"]) != activity_id
        ]
        self._save_data(data)

    def add_task_note(self, task_id: int, content: str) -> int:
        content = content.strip()
        if not content:
            raise ValueError("Note content cannot be empty.")
        data = self._load_data()
        new_id = int(data.get("next_task_note_id", 1))
        data.setdefault("task_notes", []).append({
            "id": new_id,
            "task_id": task_id,
            "timestamp": self._now_iso(),
            "user": self.current_user,
            "content": content,
        })
        data["next_task_note_id"] = new_id + 1
        self._save_data(data)
        return new_id

    def list_task_notes(self, task_id: int) -> list[TaskNoteRecord]:
        data = self._load_data()
        entries = [e for e in data.get("task_notes", []) if int(e["task_id"]) == task_id]
        entries.sort(key=lambda e: e.get("timestamp", ""))
        return [
            TaskNoteRecord(
                id=e["id"],
                task_id=e["task_id"],
                timestamp=e["timestamp"],
                user=e["user"],
                content=e["content"],
            )
            for e in entries
        ]

    def list_task_notes_for_project(self, project_id: int) -> dict[int, int]:
        """Returns {task_id: note_count} for all tasks in a project."""
        data = self._load_data()
        task_ids = {
            int(t["id"]) for t in data.get("tasks", [])
            if int(t.get("project_id", 0)) == project_id
        }
        counts: dict[int, int] = {}
        for e in data.get("task_notes", []):
            tid = int(e["task_id"])
            if tid in task_ids:
                counts[tid] = counts.get(tid, 0) + 1
        return counts

    def delete_task_note(self, note_id: int) -> None:
        data = self._load_data()
        data["task_notes"] = [
            n for n in data.get("task_notes", []) if int(n["id"]) != note_id
        ]
        self._save_data(data)

    def get_dashboard_stats(self) -> dict:
        """Summary statistics for the home screen dashboard."""
        from datetime import timedelta
        data = self._load_data()
        projects = data.get("projects", [])
        tasks = data.get("tasks", [])
        today = date.today().isoformat()
        week_str = (date.today() + timedelta(days=7)).isoformat()

        active_projects = [p for p in projects if not p.get("is_test", False)]
        incomplete_tasks = [t for t in tasks if not t.get("is_complete", False)]
        overdue = [t for t in incomplete_tasks if t.get("due_date") and t["due_date"] < today]
        due_this_week = [
            t for t in incomplete_tasks
            if t.get("due_date") and today <= t["due_date"] <= week_str
        ]

        activity = data.get("activity_log", [])
        activity_sorted = sorted(activity, key=lambda a: a.get("timestamp", ""), reverse=True)
        proj_names = {int(p["id"]): p.get("job_name", "Unknown") for p in projects}
        recent_activity = [
            {
                "timestamp": a.get("timestamp", ""),
                "user": a.get("user", ""),
                "action": a.get("action", ""),
                "entity_type": a.get("entity_type", ""),
                "entity_name": a.get("entity_name", ""),
                "project_name": proj_names.get(int(a.get("project_id", 0)), "Unknown"),
                "details": a.get("details", ""),
            }
            for a in activity_sorted[:20]
        ]

        def _proj_row(p: dict) -> dict:
            return {
                "id": int(p["id"]),
                "job_name": p.get("job_name", ""),
                "job_number": p.get("job_number", ""),
                "contract_value": p.get("contract_value", ""),
                "created_at": p.get("created_at", ""),
            }

        sortable_cv = []
        for p in active_projects:
            try:
                cv = float(p.get("contract_value") or 0)
            except (TypeError, ValueError):
                cv = 0.0
            if cv > 0:
                sortable_cv.append((cv, p))
        sortable_cv.sort(key=lambda x: x[0], reverse=True)
        top_contract = [_proj_row(p) for _, p in sortable_cv[:5]]

        newest = sorted(
            active_projects,
            key=lambda p: p.get("created_at") or "",
            reverse=True,
        )[:5]
        top_newest = [_proj_row(p) for p in newest]

        return {
            "project_count": len(active_projects),
            "overdue_count": len(overdue),
            "due_this_week_count": len(due_this_week),
            "total_tasks": len(tasks),
            "recent_activity": recent_activity,
            "top_contract": top_contract,
            "top_newest": top_newest,
        }

    # ---------- internal helpers ----------

    @staticmethod
    def _insert_default_tasks(
        data: dict[str, Any],
        project_id: int,
        task_list: tuple[dict[str, Any], ...] = DEFAULT_TASKS,
    ) -> None:
        for index, task_template in enumerate(task_list, start=1):
            new_task_id = int(data["next_task_id"])
            data["tasks"].append(
                {
                    "id": new_task_id,
                    "project_id": project_id,
                    "task_name": task_template["task_name"],
                    "phase": task_template["phase"],
                    "sort_order": index,
                    "completed_date": None,
                    "is_complete": False,
                    "notes": "",
                }
            )
            data["next_task_id"] = new_task_id + 1

    @staticmethod
    def _next_sort_order_from_data(data: dict[str, Any], project_id: int) -> int:
        matching_orders = [
            int(item.get("sort_order", 0))
            for item in data["tasks"]
            if int(item["project_id"]) == project_id
        ]
        return (max(matching_orders) + 1) if matching_orders else 1

    @staticmethod
    def _find_project_dict(
        data: dict[str, Any], project_id: int
    ) -> Optional[dict[str, Any]]:
        for project_dict in data["projects"]:
            if int(project_dict["id"]) == project_id:
                return project_dict
        return None

    @staticmethod
    def _find_task_dict(
        data: dict[str, Any], task_id: int
    ) -> Optional[dict[str, Any]]:
        for task_dict in data["tasks"]:
            if int(task_dict["id"]) == task_id:
                return task_dict
        return None

    def _extract_tasks_from_sheet(self, sheet: Any) -> list[dict[str, Any]]:
        left_rows = [10, 12, 14, 17, 20, 22, 24, 27, 29, 31, 33, 36, 38, 40, 42, 44, 51, 54]
        right_rows = [10, 12, 14, 17, 20, 22, 24, 27, 31, 33, 36, 38, 40, 42, 44, 47, 49, 51, 54]

        extracted_tasks: list[dict[str, Any]] = []
        for order_value, row_idx in enumerate(left_rows, start=1):
            task_name = self._clean_text(self._value(sheet, f"B{row_idx}"))
            if task_name:
                extracted_tasks.append(
                    {
                        "task_name": task_name,
                        "phase": self._infer_phase(task_name),
                        "sort_order": order_value,
                        "completed_date": self._value(sheet, f"D{row_idx}"),
                    }
                )

        start_order = len(extracted_tasks) + 1
        for offset, row_idx in enumerate(right_rows, start=0):
            task_name = self._clean_text(self._value(sheet, f"F{row_idx}"))
            if task_name:
                extracted_tasks.append(
                    {
                        "task_name": task_name,
                        "phase": self._infer_phase(task_name),
                        "sort_order": start_order + offset,
                        "completed_date": self._value(sheet, f"H{row_idx}"),
                    }
                )
        return extracted_tasks

    @staticmethod
    def _infer_phase(task_name: str) -> str:
        lowered = task_name.casefold()
        rules = {
            "materials": ["material", "valves", "shipping"],
            "engineering": ["drawing", "ddc", "graphics", "re-estimate", "flow curves"],
            "installation": ["elec", "install", "sub request"],
            "commissioning": ["commission", "checkout", "start-up", "service turnover"],
            "closeout": ["punch", "close-out", "owner training", "warranty"],
            "archive": ["archive", "as-built", "scan", "record drawings", "job back-up"],
            "financial": ["billing", "payment", "change-orders", "trailing costs", "database"],
            "planning": ["turnover", "review", "contract", "job plan"],
        }
        for phase_name, keywords in rules.items():
            if any(keyword in lowered for keyword in keywords):
                return phase_name.title()
        return "General"

    @staticmethod
    def _project_from_dict(project_dict: dict[str, Any]) -> ProjectRecord:
        return ProjectRecord(
            id=project_dict["id"],
            job_name=project_dict["job_name"],
            job_number=project_dict["job_number"],
            project_manager=project_dict["project_manager"],
            sales_engineer=project_dict["sales_engineer"],
            target_completion=project_dict["target_completion"],
            liquid_damages=project_dict["liquid_damages"],
            warranty_period=project_dict["warranty_period"],
            notes=project_dict["notes"],
            booked_date=project_dict.get("booked_date", ""),
            group_ops_manager=project_dict.get("group_ops_manager", ""),
            group_ops_supervisor=project_dict.get("group_ops_supervisor", ""),
            job_subtype=project_dict.get("job_subtype", ""),
            owner=project_dict.get("owner", ""),
            contracted_with=project_dict.get("contracted_with", ""),
            general_contractor=project_dict.get("general_contractor", ""),
            contract_value=project_dict.get("contract_value", ""),
            job_docs=project_dict.get("job_docs", ""),
            div25_url=project_dict.get("div25_url", ""),
            webpro_id=project_dict.get("webpro_id", ""),
            is_test=bool(project_dict.get("is_test", False)),
            pinned=bool(project_dict.get("pinned", False)),
            rss_files=_migrate_rss_files(project_dict),
            created_at=project_dict["created_at"],
            updated_at=project_dict["updated_at"],
            created_by=project_dict.get("created_by", ""),
            updated_by=project_dict.get("updated_by", ""),
        )

    @staticmethod
    def _task_from_dict(task_dict: dict[str, Any]) -> TaskRecord:
        return TaskRecord(
            id=task_dict["id"],
            project_id=task_dict["project_id"],
            task_name=task_dict["task_name"],
            phase=task_dict["phase"],
            sort_order=task_dict["sort_order"],
            due_date=task_dict.get("due_date"),
            completed_date=task_dict["completed_date"],
            is_complete=bool(task_dict["is_complete"]),
            notes=task_dict["notes"],
            updated_by=task_dict.get("updated_by", ""),
        )

    @staticmethod
    def _clean_text(value: Any) -> str:
        if value is None:
            return ""
        return " ".join(str(value).replace("\n", " ").split()).strip()

    @staticmethod
    def _value(sheet: Any, cell_ref: str) -> str:
        value = sheet[cell_ref].value
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return ProjectTrackerBackend._clean_text(value)

    @staticmethod
    def _normalize_date(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

        text = str(value).strip()
        if not text:
            return None

        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue

        # No known format matched — raise rather than silently passing through
        # a raw string that could corrupt date comparisons downstream.
        raise ValueError(
            f"Unrecognized date format: '{text}'. "
            "Expected one of: YYYY-MM-DD, MM/DD/YYYY, MM/DD/YY, YYYY/MM/DD."
        )

    @staticmethod
    def _now_iso() -> str:
        return datetime.now().replace(microsecond=0).isoformat(sep=" ")


if __name__ == "__main__":
    # Demo: creates a temporary file so it never pollutes the working directory.
    # json and tempfile are already imported at module level — no re-import needed.
    with tempfile.NamedTemporaryFile(
        suffix=".json", delete=False, mode="w"
    ) as _tmp:
        demo_path = _tmp.name

    backend = ProjectTrackerBackend(demo_path)
    demo_project_id = backend.create_project(
        ProjectRecord(
            job_name="Phoenix Demo Job",
            job_number="PHX-001",
            project_manager="Justin",
            sales_engineer="ATS",
            target_completion="2026-06-30",
        )
    )
    print(json.dumps(backend.get_project_summary(demo_project_id), indent=2))
    print(f"\nTemp data written to: {demo_path}")